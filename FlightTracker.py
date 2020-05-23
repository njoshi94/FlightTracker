# -*- coding: utf-8 -*-
"""
The following program collects flight information from 
Kayak, Google, and Expedia to to help predict when 
prices to specified places will be the lowest.

Data is cleaned and stored in SQL and results are presented with Tableau
"""

from selenium import webdriver
from datetime import date
import pandas as pd
import time
import threading
import pyodbc  
import warnings
import itertools
import matplotlib.pyplot as plt
warnings.filterwarnings("ignore")
plt.style.use('fivethirtyeight')
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def f():
    #Empty function needed for timer program 
    return

def Kayak_Function(URL, Depart_Date, Return_Date, var): 
    browser = webdriver.Chrome(executable_path='chromedriver.exe', options = chrome_options)

    #Time out if scraper gets stuck
    t1 = threading.Timer(120, f)
    t1.start()
    while t1.is_alive() == True:
        #Occassional misc errors when scraping
        #Will skip data and continue with scraping if this is case 
        try:
            #Custom url generated in SQL 
            browser.get(URL)
            time.sleep(8)
            xp_popup_close = '//button[contains(@id, "dialog-close") and contains (@class, "Button-No-Standard-Style close ")]'
            try:
                browser.find_elements_by_xpath(xp_popup_close)[8].click()
            except:
                browser.find_elements_by_xpath(xp_popup_close)[9].click()
            time.sleep(5)
        
            xp_path = '//*[contains(@src, "/res/images/horizon/flights/results/noFlightResults@2x.png")]'
            No_Flight = browser.find_elements_by_xpath(xp_path)
            i = 0 
            j = 0 
            for i in range (0, len(No_Flight)-i):
                #Checks if there is no flight data 
                if No_Flight[i].text == 'No Flights Found':
                    j = 1
                    
            if j == 0:
                #Price
                xp_path = '//*[@class = "price-text"]'
                Kayak_Price = browser.find_element_by_xpath(xp_path).text
                
                #Depart Airport 
                xp_path = '//*[contains(@class, "bottom-airport js-airport")]' 
                Kayak_Depart_Airport = browser.find_elements_by_xpath(xp_path)[0].text
                
                #Depart Airline 
                xp_path = '//*[@class = "bottom "]'                
                Kayak_Depart_Airline = browser.find_elements_by_xpath(xp_path)[0].text
               
                #Direct / Stops 
                xp_path = '//*[@class = "section stops"]'                

                Kayak_Depart_Stops = browser.find_elements_by_xpath(xp_path)[0].text
                Kayak_Depart_Stops = Kayak_Depart_Stops.splitlines()[0]

                #Total time
                xp_path = '//*[@class = "top"]'
                Kayak_Depart_Total_Time = browser.find_elements_by_xpath(xp_path)[2].text
                
                #Depart Time
                xp_path = '//*[@class = "depart-time base-time"]'

                Kayak_Depart_Time1 = browser.find_element_by_xpath(xp_path).text
                xp_path = '//*[@class = "time-meridiem meridiem"]'
                Kayak_Depart_Time2 = browser.find_element_by_xpath(xp_path).text
                Kayak_Depart_Time = Kayak_Depart_Time1 + Kayak_Depart_Time2
        
                #Return Airport 
                xp_path = '//*[contains(@class, "bottom-airport js-airport")]'
                Kayak_Return_Airport = browser.find_elements_by_xpath(xp_path)[1].text
        
                #Return Airline 
                xp_path = '//*[contains(@class, "bottom ")]'
                Kayak_Return_Airline = browser.find_elements_by_xpath(xp_path)[3].text

                #Direct / Stops 
                xp_path = '//*[@class = "section stops"]'

                Kayak_Return_Stops = browser.find_elements_by_xpath(xp_path)[1].text
                Kayak_Return_Stops = Kayak_Return_Stops.splitlines()[0]

                #Total time 
                xp_path = '//*[@class = "top"]'

                Kayak_Return_Total_Time = browser.find_elements_by_xpath(xp_path)[5].text

                #Return Time
                xp_path = '//*[@class = "depart-time base-time"]'
                
                Kayak_Return_Time1 = browser.find_elements_by_xpath(xp_path)[1].text
                xp_path = '//*[@class = "time-meridiem meridiem"]'
                Kayak_Return_Time2 = browser.find_elements_by_xpath(xp_path)[1].text
                Kayak_Return_Time = Kayak_Return_Time1 + Kayak_Return_Time2
                
                #Inserts the results into the correct table 
                if var == 0:
                    insert_query = '''INSERT INTO Python_Precleaning
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
                else:
                    insert_query = '''INSERT INTO Python_Precleaning_Standard
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''            
                Values = (date.today(), Depart_Date, Return_Date,'Kayak', Kayak_Price, Kayak_Depart_Airport,
                          Kayak_Depart_Airline, Kayak_Depart_Stops, Kayak_Depart_Total_Time,
                          Kayak_Depart_Time, Kayak_Return_Airport,
                          Kayak_Return_Airline, Kayak_Return_Stops, Kayak_Return_Total_Time,
                          Kayak_Return_Time, URL)
                cursor.execute(insert_query, Values)
                conn.commit()
            browser.quit()
        except:
            pass
        t1.cancel()
    
def Google_Function(URL, Depart_Date, Return_Date, var):
    #See comments from Kayak Function
    browser = webdriver.Chrome(executable_path='chromedriver.exe', options = chrome_options)
    t1 = threading.Timer(180, f)
    t1.start()
    while t1.is_alive() == True:
        try:
            browser.get(URL)
            time.sleep(5)
            Error = True 
            Iter = 0 
            
            while Error and Iter < 5:            
                try:
                    xp_path = '//*[@class = "gws-flights-results__error-page gws-flights__center-content"]'
                    browser.find_element_by_xpath(xp_path).text
                    browser.refresh()
                    time.sleep(5)
                    Error = True 
                except:
                    Error = False
                Iter = Iter + 1 
            
            if Error == False:
                try:
                    xp_path = '//*[@class = "gws-flights__flex-grow gws-flights-results__results gws-flights__flex-column gws-flights__scrollbar-padding gws-flights-results__error"]'
                    browser.find_element_by_xpath(xp_path).text
                except:
                    #Price
                    xp_path = '//*[@class = "flt-subhead1 gws-flights-results__price gws-flights-results__cheapest-price"]'
                    Google_Price = browser.find_element_by_xpath(xp_path).text
                    
                    #Depart Airport 
                    xp_path = '//*[contains(@class, "gws-flights-results__airports flt-caption")]' 
                    Google_Depart_Airport = browser.find_elements_by_xpath(xp_path)[0].text
                    Google_Depart_Airport = Google_Depart_Airport.split('–')[0]
                    
                    #Depart Airline 
                    xp_path = '//*[@class = "gws-flights__ellipsize"]'
                    Google_Depart_Airline = browser.find_elements_by_xpath(xp_path)[0].text
                        
                    #Direct / Stops 
                    xp_path = '//*[contains(@class, "gws-flights-results__stops flt-subhead1Normal")]'
                    Google_Depart_Stops = browser.find_elements_by_xpath(xp_path)[0].text
                    
                    #Total time 
                    xp_path = '//*[@class = "gws-flights-results__duration flt-subhead1Normal"]'
                    Google_Depart_Total_Time = browser.find_elements_by_xpath(xp_path)[0].text
                    
                    #Depart Time
                    xp_path = '//*[@class = "gws-flights-results__times flt-subhead1"]'
                    Google_Depart_Time = browser.find_element_by_xpath(xp_path).text
                    Google_Depart_Time = Google_Depart_Time.split('–')[0]
                    
                    time.sleep(5)
                    
                    xp_popup_close =  '//*[@class = "gws-flights-results__collapsed-itinerary gws-flights-results__itinerary gws-flights-results__itinerary-narrow"]'
                    browser.find_elements_by_xpath(xp_popup_close)[0].click()
                    
                    time.sleep(10)
                    
                    Error = True 
                    Iter = 0 
                    while Error and Iter < 5:            
                        try:
                            xp_path = '//*[@class = "gws-flights-results__error-page gws-flights__center-content"]'
                            browser.find_element_by_xpath(xp_path).text
                            browser.refresh()
                            time.sleep(5)
                            Error = True 
                        except:
                            Error = False
                        Iter = Iter + 1 
                    
                    if Error == False:
                        try:
                            xp_path = '//*[@class = "gws-flights__flex-grow gws-flights-results__results gws-flights__flex-column gws-flights__scrollbar-padding gws-flights-results__error"]'
                            browser.find_element_by_xpath(xp_path).text
                        except:
                            #Return Airport 
                            xp_path = '//*[contains(@class, "gws-flights-results__airports flt-caption")]' 
                            Google_Return_Airport = browser.find_elements_by_xpath(xp_path)[0].text
                            Google_Return_Airport = Google_Return_Airport.split('–')[0]
                            
                            #Depart Airline 
                            xp_path = '//*[@class = "gws-flights__ellipsize"]'
                            Google_Return_Airline = browser.find_elements_by_xpath(xp_path)[0].text
                            
                            #Direct / Stops 
                            xp_path = '//*[contains(@class, "gws-flights-results__stops flt-subhead1Normal")]'
                            Google_Return_Stops = browser.find_elements_by_xpath(xp_path)[0].text
                            
                            #Total time 
                            xp_path = '//*[@class = "gws-flights-results__duration flt-subhead1Normal"]'
                            Google_Return_Total_Time = browser.find_elements_by_xpath(xp_path)[0].text
                            
                            #Depart Time
                            xp_path = '//*[@class = "gws-flights-results__times flt-subhead1"]'
                            Google_Return_Time = browser.find_element_by_xpath(xp_path).text
                            Google_Return_Time = Google_Return_Time.split('–')[0]
                            
                            if var == 0:
                                insert_query = '''INSERT INTO Python_Precleaning
                                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
                            else:
                                insert_query = '''INSERT INTO Python_Precleaning_Standard
                                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''                        
                            Values = (date.today(), Depart_Date, Return_Date,'Google', Google_Price, Google_Depart_Airport,
                                      Google_Depart_Airline, Google_Depart_Stops, Google_Depart_Total_Time,
                                      Google_Depart_Time, Google_Return_Airport,
                                      Google_Return_Airline, Google_Return_Stops, Google_Return_Total_Time,
                                      Google_Return_Time, URL)
                            cursor.execute(insert_query, Values)
                            conn.commit()
            browser.quit()
        except:
            pass
        t1.cancel()

def Expedia_Function(URL,Depart_Date, Return_Date, var):
   #See comments from Kayak Function
    browser = webdriver.Chrome(executable_path='chromedriver.exe', options = chrome_options)
    t1 = threading.Timer(120, f)
    t1.start()
    while t1.is_alive() == True:
        try:
            browser.get(URL)
            time.sleep(7)
            try:
                xp_path = '//*[@class = "bold announce-able"]'
                browser.find_element_by_xpath(xp_path).text
            except:
                #Price
                xp_path = '//*[@class = "full-bold no-wrap"]'
                Expedia_Price = browser.find_element_by_xpath(xp_path).text
                
                #Depart Airport 
                xp_path = '//*[contains(@class, "secondary-content no-wrap")]' 
                Expedia_Depart_Airport = browser.find_elements_by_xpath(xp_path)[0].text
                Expedia_Depart_Airport = Expedia_Depart_Airport.splitlines()[1].split(' - ')[0]
                
                #Depart Airline 
                xp_path = '//*[@data-test-id = "airline-name"]'
                Expedia_Depart_Airline = browser.find_elements_by_xpath(xp_path)[1].text
                
                
                #Direct / Stops 
                xp_path = '//*[@class = "number-stops"]'
                Expedia_Depart_Stops = browser.find_elements_by_xpath(xp_path)[0].text
                
                #Total time 
                xp_path = '//*[@class = "duration-emphasis"]'
                Expedia_Depart_Total_Time = browser.find_elements_by_xpath(xp_path)[0].text
                
                #Depart Time
                xp_path = '//*[@data-test-id = "departure-time"]'
                Expedia_Depart_Time = browser.find_element_by_xpath(xp_path).text
                
                xp_popup_close = '//button[contains (@class, "btn-secondary btn-action t-select-btn")]'
                
                browser.find_elements_by_xpath(xp_popup_close)[0].click()
                
                time.sleep(5)
                
                browser.find_elements_by_xpath(xp_popup_close)[1].click()
                
                time.sleep(7)
                
                #Return Airport 
                xp_path = '//*[contains(@class, "secondary-content no-wrap")]' 
                Expedia_Return_Airport = browser.find_elements_by_xpath(xp_path)[0].text
                Expedia_Return_Airport = Expedia_Return_Airport.splitlines()[1].split(' - ')[0]
                
                #Return Airline 
                xp_path = '//*[@data-test-id = "airline-name"]'
                Expedia_Return_Airline = browser.find_elements_by_xpath(xp_path)[2].text
                
                #Direct / Stops 
                xp_path = '//*[@class = "number-stops"]'
                Expedia_Return_Stops = browser.find_elements_by_xpath(xp_path)[1].text
                
                #Total time 
                xp_path = '//*[@class = "duration-emphasis"]'
                Expedia_Return_Total_Time = browser.find_elements_by_xpath(xp_path)[1].text
                
                #Return Time
                xp_path = '//*[@data-test-id = "departure-time"]'
                Expedia_Return_Time = browser.find_elements_by_xpath(xp_path)[1].text
                
                if var == 0:
                    insert_query = '''INSERT INTO Python_Precleaning
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
                else:
                    insert_query = '''INSERT INTO Python_Precleaning_Standard
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
                Values = (date.today(), Depart_Date, Return_Date,'Expedia', Expedia_Price, Expedia_Depart_Airport,
                          Expedia_Depart_Airline, Expedia_Depart_Stops, Expedia_Depart_Total_Time,
                          Expedia_Depart_Time, Expedia_Return_Airport,
                          Expedia_Return_Airline, Expedia_Return_Stops, Expedia_Return_Total_Time,
                          Expedia_Return_Time,URL) 
                cursor.execute(insert_query, Values)
                conn.commit()
            browser.quit()
        except:
            pass
        t1.cancel()

def ML_Predictions(SQL_Query, IATA, val):    
    #Function to predict lowest prices for following two weeks
    
    #Adds data for any days that are missing by using 
    #previous day's lowest price
    for row in SQL_Query['date']:
        if row + pd.Timedelta(days = 1) not in SQL_Query['date'].values and row != max(SQL_Query['date']):
            pricer = SQL_Query.loc[SQL_Query['date'] == row, 'price']
            data = pd.DataFrame({
            'date': [row + pd.Timedelta(days = 1)],
            'price': [min(pricer)]
            })
            SQL_Query = SQL_Query.append(data)
    
    #If less than 14 days, will backfill to generate predictions
    while len(SQL_Query) < 14:
        Day = min(SQL_Query['date']) - pd.Timedelta(days = 1)
        pricer = SQL_Query.loc[SQL_Query['date'] == min(SQL_Query['date']), 'price']
        data = pd.DataFrame({
            'date': [Day],
            'price': [min(pricer)]
            })
        SQL_Query = SQL_Query.append(data)
    
    SQL_Query = SQL_Query.sort_values(by = 'date')
    
    SQL_Query['date'] = pd.to_datetime(SQL_Query['date'])
    
    df = SQL_Query.set_index(pd.DatetimeIndex(SQL_Query['date']))
    df.index = pd.DatetimeIndex(df.index.values,
                                   freq=df.index.inferred_freq)
    
    y = pd.DataFrame(df['price'])
    
    p = d = q = range(0, 2)
    pdq = list(itertools.product(p, d, q))
    seasonal_pdq = [(x[0], x[1], x[2], 12) for x in list(itertools.product(p, d, q))]
    
    ord = ()
    seas_ord = ()
    AIC = 500  
    
    #Finds the optical pdq and seasonal pdq values using AIC 
    for param in pdq:
        for param_seasonal in seasonal_pdq:
            try:
                mod = sm.tsa.statespace.SARIMAX(y,
                order=param,
                seasonal_order=param_seasonal,
                enforce_stationarity=False,
                enforce_invertibility=False)
                results = mod.fit()
                if results.aic < AIC:
                    AIC = results.aic 
                    ord = param
                    seas_ord = param_seasonal      
            except:
                continue
    
    
    mod = sm.tsa.statespace.SARIMAX(y,
                                order=ord,
                                seasonal_order= seas_ord ,
                                enforce_stationarity=False,
                                enforce_invertibility=False)
    results = mod.fit()
    
    #This value is strictly for viewing the models in Python
    z = min(y.index)
    
    #Forecasting two weeks in advance 
    pred = results.get_forecast(steps = 14)
    pred_ci = pred.conf_int()
    ax = y[z:].plot(label='observed')
    pred.predicted_mean.plot(ax=ax, label='One-step ahead Forecast', alpha=.7, figsize=(14, 7))
    ax.fill_between(pred_ci.index,
                    pred_ci.iloc[:, 0],
                    pred_ci.iloc[:, 1], color='blue', alpha=.2)

    ax.set_xlabel('Date')
    ax.set_ylabel('Ticket Prices')
    plt.legend()
    plt.show()
    
    #Adds values to correct table     
    if val == 0:
        insert_query = '''INSERT INTO Python_Preds
        VALUES(?,?,?,?)'''
    else:
        insert_query = '''INSERT INTO Python_Preds_Standard
        VALUES(?,?,?,?)'''
    i = 0 
    while i < len(pred.predicted_mean):       
        Values = (date.today(), IATA, pred.predicted_mean.index[i], pred.predicted_mean[i] ) 
        cursor.execute(insert_query, Values)
        conn.commit()
        i = i + 1 

#Criteria to connect to browser
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")

#Criteria to connect to SQL 
conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=DESKTOP-A7U94PB;'
                      'Database=master;'
                      'Trusted_Connection=yes;')

#Generates list of URLS for scraping 
cursor = conn.cursor()
cursor.execute('exec PRECLEANING_PROCEDURE')
time.sleep(10)
conn.commit()

SQL_Query = pd.read_sql_query(
    '''SELECT * FROM master.dbo.custom_url_list''', conn)

#Iterates through all items in list
i = 0 
var = 0
while i < len(SQL_Query):
     Kayak_Function(SQL_Query['Kayak_Link'][i], SQL_Query['Depart_Date'][i], SQL_Query['Return_Date'][i], var)
     Google_Function(SQL_Query['Google_Link'][i], SQL_Query['Depart_Date'][i], SQL_Query['Return_Date'][i], var)
     Expedia_Function(SQL_Query['Expedia_Link'][i], SQL_Query['Depart_Date'][i], SQL_Query['Return_Date'][i], var)
     i = i + 1 


#Generates list of URLs in Standard Table 
SQL_Query = pd.read_sql_query(
    '''SELECT * FROM master.dbo.standard_url_list''', conn)

#Iterates through all items in list 
i = 0 
var = 1
while i < len(SQL_Query):
     Kayak_Function(SQL_Query['Kayak_Link'][i], SQL_Query['Depart_Date'][i], SQL_Query['Return_Date'][i], var)
     Google_Function(SQL_Query['Google_Link'][i], SQL_Query['Depart_Date'][i], SQL_Query['Return_Date'][i], var)
     Expedia_Function(SQL_Query['Expedia_Link'][i], SQL_Query['Depart_Date'][i], SQL_Query['Return_Date'][i], var)
     i = i + 1
 
    
#Cleans all the new data 
cursor.execute('exec PYTHON_CLEANING')
time.sleep(15)
conn.commit()

#Iterates through all destinations for predictions 
VALS = pd.read_sql_query(
    """select distinct return_airport 
from Python_Cleaned
""", conn)
conn.commit()

i = 0 
while i < len(VALS):
        SQL_Query = pd.read_sql_query(
        """select [date], min(price) as price 
        from Python_Cleaned
        where return_airport = '""" + VALS['return_airport'][i] + """' 
        group by [date] 
        """, conn)
        conn.commit()
    
        ML_Predictions(SQL_Query, VALS['return_airport'][i], 0)
        i = i + 1 

VALS = pd.read_sql_query(
    """select distinct return_airport 
from Python_Cleaned_Standard
""", conn)
conn.commit()

i = 0 
while i < len(VALS):
        SQL_Query = pd.read_sql_query(
        """select [date], min(price) as price 
        from Python_Cleaned_standard
        where return_airport = '""" + VALS['return_airport'][i] + """' 
        group by [date] 
        """, conn)
        conn.commit()
        ML_Predictions(SQL_Query, VALS['return_airport'][i], 1)
        i = i + 1 
        
#Creates final table with newest predictions 
cursor.execute('exec CREATE_FINAL_TABLE')
time.sleep(15)
conn.commit()        

#Updates Excel file for Tableau
#Only required for free version of Tableau 
SQL_Query = pd.read_sql_query(
"""select *  
from Flight_Data_Final
""", conn)
conn.commit()

filename = 'Flight_Data.xlsx'
wb = load_workbook(filename)
delt = wb['Sheet']
wb.remove(delt)
ws2 = wb.create_sheet('Sheet')
for r in dataframe_to_rows(SQL_Query, index=False, header=True):
    ws2.append(r)
wb.save(filename)
